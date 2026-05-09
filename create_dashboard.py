import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
CSV_FILE = BASE_DIR / "processed_users.csv"
OUTPUT_FILE = BASE_DIR / "dashboard.xlsx"


def read_rows() -> list[dict]:
    if not CSV_FILE.exists():
        raise FileNotFoundError(f"Expected CSV at {CSV_FILE}")
    with CSV_FILE.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def autosize_columns(sheet) -> None:
    for col_cells in sheet.columns:
        max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
        sheet.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2


def build_dashboard() -> None:
    rows = read_rows()
    if not rows:
        raise ValueError("No data found in processed_users.csv")

    wb = Workbook()
    data_sheet = wb.active or wb.create_sheet()
    data_sheet.title = "Processed Data"

    headers = ["user_id", "name", "email", "city", "company", "total_posts"]
    data_sheet.append(headers)
    for row in rows:
        data_sheet.append([row[key] for key in headers])

    dash = wb.create_sheet("Dashboard")
    dash["A1"] = "Automation Data Integration Dashboard"
    dash["A1"].font = Font(size=16, bold=True, color="1F4E78")

    avg_posts = sum(int(r["total_posts"]) for r in rows) / len(rows)
    dash["A3"] = "Average posts per user"
    dash["B3"] = round(avg_posts, 2)
    dash["B3"].number_format = "0.00"

    users_by_city: dict[str, int] = defaultdict(int)
    posts_by_company: dict[str, int] = defaultdict(int)
    for row in rows:
        users_by_city[row["city"]] += 1
        posts_by_company[row["company"]] += int(row["total_posts"])

    for cell_ref, label in (("A5", "Users per City"), ("D5", "Posts per Company")):
        dash[cell_ref] = label
        dash[cell_ref].font = Font(bold=True, color="FFFFFF")
        dash[cell_ref].fill = PatternFill("solid", fgColor="1F4E78")

    dash.append(["City", "Users"])
    for city, count in sorted(users_by_city.items()):
        dash.append([city, count])

    header_row = 6
    dash.cell(row=header_row, column=4, value="Company")
    dash.cell(row=header_row, column=5, value="Posts")
    for i, (company, total) in enumerate(sorted(posts_by_company.items()), start=1):
        dash.cell(row=header_row + i, column=4, value=company)
        dash.cell(row=header_row + i, column=5, value=total)

    city_chart = BarChart()
    city_chart.title = "Users per City"
    city_chart.y_axis.title = "Users"
    city_chart.x_axis.title = "City"
    city_data = Reference(dash, min_col=2, min_row=6, max_row=6 + len(users_by_city))
    city_labels = Reference(dash, min_col=1, min_row=7, max_row=6 + len(users_by_city))
    city_chart.add_data(city_data, titles_from_data=True)
    city_chart.set_categories(city_labels)
    dash.add_chart(city_chart, "A19")

    company_chart = BarChart()
    company_chart.title = "Posts per Company"
    company_chart.y_axis.title = "Posts"
    company_chart.x_axis.title = "Company"
    company_data = Reference(dash, min_col=5, min_row=6, max_row=6 + len(posts_by_company))
    company_labels = Reference(dash, min_col=4, min_row=7, max_row=6 + len(posts_by_company))
    company_chart.add_data(company_data, titles_from_data=True)
    company_chart.set_categories(company_labels)
    dash.add_chart(company_chart, "J19")

    for sheet in wb.worksheets:
        autosize_columns(sheet)

    wb.save(OUTPUT_FILE)
    print(f"Dashboard saved to {OUTPUT_FILE.name}")


if __name__ == "__main__":
    build_dashboard()
